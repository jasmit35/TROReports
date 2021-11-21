import logging as log
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import pathlib


class ExcelSheet:
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.workbook.iso_dates = True
        self.sheet = self.workbook.active
        self.row = 1
        self.column = 1

    def set_row_values(self, column_values):
        self.column = 1
        for this_value in column_values:
            self.sheet.cell(self.row, self.column).value = this_value
            self.column += 1

    def add_border_row(self, start_col, end_col):
        summary_border = Border(top=Side(style='double'))
        for col in range(start_col, end_col + 1):
            self.sheet.cell(self.row, col).border = summary_border

    def add_summary_cell(self, sum_start_row, sum_end_row):
        start_cell = f"{get_column_letter(self.col)}{sum_start_row}"
        end_cell = f"{get_column_letter(self.col)}{sum_end_row}"
        formula = f'=SUM({start_cell}:{end_cell})'
        self.sheet.cell(self.row, self.col).value = formula

    def format_cell_date(self):
        self.sheet.cell(self.row, self.col).number_format = "mm/dd/yy"

    def format_cell_numeric(self):
        self.sheet.cell(self.row, self.col).number_format = "#,##0.00_);[Red](#,##0.00)"

    def save(self, file_name, create_if_missing=True):
        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(self.sheet.max_column))
        for column_letter in column_letters:
            self.sheet.column_dimensions[column_letter].bestFit = True

        path = pathlib.Path(file_name)
        dir_spec = path.parent
        if (dir_spec.exists() is False) and (create_if_missing is True):
            log.info(f"Creating directory {dir_spec}")
            dir_spec.mkdir(parents=True, exist_ok=True)
        self.workbook.save(path)
